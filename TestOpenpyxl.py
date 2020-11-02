from openpyxl import load_workbook

RutaDeLaBase = "Alquileres.xlsx"
HojaDeLaBase = "Data"

workbook = load_workbook(RutaDeLaBase, read_only=True)
sheet = workbook[HojaDeLaBase]

for Dato in sheet.iter_rows():
    print(Dato[0].value,"   |   ",Dato[1].value,"   |   ",Dato[2].value,"   |   ",Dato[3].value,"   |   ",Dato[4].value,"   |   ",Dato[5].value)


